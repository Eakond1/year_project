import time
import openpyxl
import telebot
from telebot import types
import json
import datetime
import schedule

bot = telebot.TeleBot('')
users = []

try:
    with open('users.json', 'r', encoding='utf-8') as file:
        users = json.loads(file.read())
except Exception as e:
    print(e)


def save_data():
    with open('users.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(users, ensure_ascii=False))


@bot.message_handler(commands=['start', 'help'])
def start(message):
    chat_id = message.chat.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    class_10t = types.KeyboardButton("10Т")
    class_10i = types.KeyboardButton("10И")
    markup.add(class_10t, class_10i)
    m = bot.send_message(chat_id, 'Привет! В каком ты классе?', reply_markup=markup)
    bot.register_next_step_handler(m, know_user_class)


def know_user_class(message):
    chat_id = message.chat.id
    user_class = message.text
    u = find_user(chat_id)
    if u == True:
        change_user_class(chat_id, user_class)
        bot.send_message(chat_id, 'Вы успешно изменили свой класс, для начала работы бота нажмите /schedule', reply_markup=types.ReplyKeyboardRemove())
    else:
        if user_class == "10Т" or user_class == "10И":
            users.append({
                "id": chat_id,
                "class": user_class
            })
            save_data()
            bot.send_message(chat_id, 'Вы успешно завершили регистрацию, для начала работы бота введите /schedule ', reply_markup=types.ReplyKeyboardRemove())
        else:
            m = bot.send_message(chat_id, 'Повторите попытку')
            bot.register_next_step_handler(m, know_user_class)


@bot.message_handler(commands=['schedule'])
def start(message):
    chat_id = message.chat.id
    schedule.every().monday.at("08:30").do(send_notification, chat_id)
    schedule.every().monday.at("09:15").do(send_day_second, chat_id)
    schedule.every().monday.at("10:10").do(send_day_third, chat_id)
    schedule.every().monday.at("11:10").do(send_day_forth, chat_id)
    schedule.every().monday.at("12:10").do(send_day_fifth, chat_id)
    schedule.every().monday.at("13:05").do(send_day_sixth, chat_id)
    schedule.every().monday.at("14:10").do(send_day_seventh, chat_id)
    schedule.every().monday.at("15:15").do(send_day_eighth, chat_id)

    schedule.every().tuesday.at("08:30").do(send_notification, chat_id)
    schedule.every().tuesday.at("09:15").do(send_day_second, chat_id)
    schedule.every().tuesday.at("10:10").do(send_day_third, chat_id)
    schedule.every().tuesday.at("11:10").do(send_day_forth, chat_id)
    schedule.every().tuesday.at("12:10").do(send_day_fifth, chat_id)
    schedule.every().tuesday.at("13:05").do(send_day_sixth, chat_id)
    schedule.every().tuesday.at("14:10").do(send_day_seventh, chat_id)
    schedule.every().tuesday.at("15:15").do(send_day_eighth, chat_id)

    schedule.every().wednesday.at("08:30").do(send_notification, chat_id)
    schedule.every().wednesday.at("09:15").do(send_day_second, chat_id)
    schedule.every().wednesday.at("10:10").do(send_day_third, chat_id)
    schedule.every().wednesday.at("11:10").do(send_day_forth, chat_id)
    schedule.every().wednesday.at("12:10").do(send_day_fifth, chat_id)
    schedule.every().wednesday.at("13:05").do(send_day_sixth, chat_id)
    schedule.every().wednesday.at("14:10").do(send_day_seventh, chat_id)

    schedule.every().thursday.at("08:30").do(send_notification, chat_id)
    schedule.every().thursday.at("09:15").do(send_day_second, chat_id)
    schedule.every().thursday.at("10:10").do(send_day_third, chat_id)
    schedule.every().thursday.at("11:10").do(send_day_forth, chat_id)
    schedule.every().thursday.at("12:10").do(send_day_fifth, chat_id)
    schedule.every().thursday.at("13:05").do(send_day_sixth, chat_id)
    schedule.every().thursday.at("14:10").do(send_day_seventh, chat_id)

    schedule.every().friday.at("08:30").do(send_notification, chat_id)
    schedule.every().friday.at("09:15").do(send_day_second, chat_id)
    schedule.every().friday.at("10:10").do(send_day_third, chat_id)
    schedule.every().friday.at("11:10").do(send_day_forth, chat_id)
    schedule.every().friday.at("12:10").do(send_day_fifth, chat_id)
    schedule.every().friday.at("13:05").do(send_day_sixth, chat_id)
    schedule.every().friday.at("14:10").do(send_day_seventh, chat_id)


def send_lesson_10t(number):
    wd = datetime.datetime.today().weekday()
    if wd == 0:
        dataframe = openpyxl.load_workbook("расписание на понедельник.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{8}'].value
            name2 = active[f'C{9}'].value
            cabinet = active[f'A{8}'].value
            cabinet2 = active[f'A{9}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};{name2}, каб. {cabinet2} ')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{8}'].value
            name2 = active[f'D{9}'].value
            cabinet = active[f'A{8}'].value
            cabinet2 = active[f'A{9}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};{name2}, каб. {cabinet2} ')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{8}'].value
            name2 = active[f'E{9}'].value
            cabinet = active[f'A{8}'].value
            cabinet2 = active[f'A{9}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};{name2}, каб. {cabinet2} ')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{8}'].value
            name2 = active[f'F{9}'].value
            cabinet = active[f'A{8}'].value
            cabinet2 = active[f'A{9}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};{name2}, каб. {cabinet2} ')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 8:
            num = active[f'I{1}'].value
            name1 = active[f'I{3}'].value
            name2 = active[f'I{4}'].value
            cabinet = active[f'A{3}'].value
            cabinet2 = active[f'A{4}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};{name2}, каб. {cabinet2} ')
            return lesson
    elif wd == 1:
        dataframe = openpyxl.load_workbook("расписание на вторник.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{14}'].value
            cabinet = active[f'A{14}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{16}'].value
            cabinet = active[f'A{16}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{8}'].value
            cabinet = active[f'A{8}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{4}'].value
            cabinet = active[f'A{4}'].value
            name2 = active[f'G{3}'].value
            cabinet2 = active[f'A{3}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}, {name2}, каб. {cabinet2}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{4}'].value
            cabinet = active[f'A{4}'].value
            name2 = active[f'G{3}'].value
            cabinet2 = active[f'A{3}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}; {name2}, каб. {cabinet2}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{4}'].value
            cabinet = active[f'A{4}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
    elif wd == 2:
        dataframe = openpyxl.load_workbook("расписание на среду.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{5}'].value
            cabinet = active[f'A{5}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{10}'].value
            cabinet = active[f'A{10}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{8}'].value
            cabinet = active[f'A{8}'].value
            name2 = active[f'E{9}'].value
            cabinet2 = active[f'A{9}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}; {name2}, каб. {cabinet2}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{8}'].value
            cabinet = active[f'A{8}'].value
            name2 = active[f'F{9}'].value
            cabinet2 = active[f'A{9}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}, {name2}, каб. {cabinet2}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
    elif wd == 3:
        dataframe = openpyxl.load_workbook("расписание на четверг.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{14}'].value
            cabinet = active[f'A{14}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{10}'].value
            cabinet = active[f'A{10}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{10}'].value
            cabinet = active[f'A{10}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{5}'].value
            cabinet = active[f'A{5}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
    elif wd == 4:
        dataframe = openpyxl.load_workbook("расписание на пятницу.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{2}'].value
            cabinet = active[f'A{2}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{2}'].value
            cabinet = active[f'A{2}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{5}'].value
            cabinet = active[f'A{5}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{5}'].value
            cabinet = active[f'A{5}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson


def send_lesson_10i(number):
    wd = datetime.datetime.today().weekday()
    if wd == 0:
        dataframe = openpyxl.load_workbook("расписание на понедельник.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{12}'].value
            cabinet = active[f'A{12}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{18}'].value
            cabinet = active[f'A{18}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{14}'].value
            cabinet = active[f'A{14}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{15}'].value
            cabinet = active[f'A{15}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{15}'].value
            cabinet = active[f'A{15}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
    elif wd == 1:
        dataframe = openpyxl.load_workbook("расписание на вторник.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{10}'].value
            cabinet = active[f'A{10}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{16}'].value
            cabinet = active[f'A{16}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{11}'].value
            cabinet = active[f'A{11}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{14}'].value
            cabinet = active[f'A{14}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{12}'].value
            cabinet = active[f'A{12}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 8:
            num = active[f'I{1}'].value
            name1 = active[f'I{3}'].value
            cabinet = active[f'A{3}'].value
            name2 = active[f'I{4}'].value
            cabinet2 = active[f'A{4}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}; {name2}, каб. {cabinet2}')
            return lesson
    elif wd == 2:
        dataframe = openpyxl.load_workbook("расписание на среду.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{12}'].value
            cabinet = active[f'A{12}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{14}'].value
            cabinet = active[f'A{14}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{3}'].value
            cabinet = active[f'A{3}'].value
            name2 = active[f'E{4}'].value
            cabinet2 = active[f'A{4}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}; {name2}, каб. {cabinet2}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{3}'].value
            cabinet = active[f'A{3}'].value
            name2 = active[f'F{4}'].value
            cabinet2 = active[f'A{4}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}, {name2}, каб. {cabinet2}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{10}'].value
            cabinet = active[f'A{10}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{10}'].value
            cabinet = active[f'A{10}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
    elif wd == 3:
        dataframe = openpyxl.load_workbook("расписание на четверг.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{12}'].value
            cabinet = active[f'A{12}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{12}'].value
            cabinet = active[f'A{12}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{9}'].value
            cabinet = active[f'A{9}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{9}'].value
            cabinet = active[f'A{9}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
    elif wd == 4:
        dataframe = openpyxl.load_workbook("расписание на пятницу.xlsx")
        active = dataframe.active
        lesson = []
        if number == 2:
            num = active[f'C{1}'].value
            name1 = active[f'C{14}'].value
            cabinet = active[f'A{14}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 3:
            num = active[f'D{1}'].value
            name1 = active[f'D{8}'].value
            cabinet = active[f'A{8}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet};')
            return lesson
        elif number == 4:
            num = active[f'E{1}'].value
            name1 = active[f'E{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 5:
            num = active[f'F{1}'].value
            name1 = active[f'F{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 6:
            num = active[f'G{1}'].value
            name1 = active[f'G{11}'].value
            cabinet = active[f'A{11}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson
        elif number == 7:
            num = active[f'H{1}'].value
            name1 = active[f'H{13}'].value
            cabinet = active[f'A{13}'].value
            lesson.append(f'{num}. {name1}, каб. {cabinet}')
            return lesson


def send_lessons(lesson, chat_id):
    for i in range(len(lesson)):
        bot.send_message(chat_id, lesson[i])
        time.sleep(1)


def get_list_of_lessons(user_class):
    wd = datetime.datetime.today().weekday()
    if user_class == "10Т":
        if wd == 0:
            dataframe = openpyxl.load_workbook("расписание на понедельник.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{4}'].value
            cabinet1 = active[f'A{4}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{8}'].value
            name3 = active[f'C{9}'].value
            cabinet2 = active[f'A{8}'].value
            cabinet3 = active[f'A{9}'].value
            num4 = active[f'D{1}'].value
            name4 = active[f'D{8}'].value
            name5 = active[f'D{9}'].value
            cabinet4 = active[f'A{8}'].value
            cabinet5 = active[f'A{9}'].value
            num6 = active[f'E{1}'].value
            name6 = active[f'E{8}'].value
            name7 = active[f'E{9}'].value
            cabinet6 = active[f'A{8}'].value
            cabinet7 = active[f'A{9}'].value
            num8 = active[f'F{1}'].value
            name8 = active[f'F{8}'].value
            name9 = active[f'F{9}'].value
            cabinet8 = active[f'A{8}'].value
            cabinet9 = active[f'A{9}'].value
            num10 = active[f'G{1}'].value
            name10 = active[f'G{13}'].value
            cabinet10 = active[f'A{13}'].value
            num11 = active[f'H{1}'].value
            name11 = active[f'H{13}'].value
            cabinet11 = active[f'A{13}'].value
            num12 = active[f'I{1}'].value
            name12 = active[f'I{3}'].value
            name13 = active[f'I{4}'].value
            cabinet12 = active[f'A{3}'].value
            cabinet13 = active[f'A{4}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num2}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num4}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num6}. {name7}, каб. {cabinet7};\n \
                            {num8}. {name8}, каб. {cabinet8};\n \
                            {num8}. {name9}, каб. {cabinet9};\n \
                            {num10}. {name10}, каб. {cabinet10};\n \
                            {num11}. {name11}, каб. {cabinet11};\n \
                            {num12}. {name12}, каб. {cabinet12};\n \
                            {num12}. {name13}, каб. {cabinet13};')
            return lesson
        elif wd == 1:
            dataframe = openpyxl.load_workbook("расписание на вторник.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{11}'].value
            cabinet1 = active[f'A{11}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{14}'].value
            cabinet2 = active[f'A{14}'].value
            num3 = active[f'D{1}'].value
            name3 = active[f'D{16}'].value
            cabinet3 = active[f'A{16}'].value
            num4 = active[f'E{1}'].value
            name4 = active[f'E{8}'].value
            cabinet4 = active[f'A{8}'].value
            num5 = active[f'F{1}'].value
            name5 = active[f'F{4}'].value
            cabinet5 = active[f'A{4}'].value
            name6 = active[f'F{3}'].value
            cabinet6 = active[f'A{3}'].value
            num7 = active[f'G{1}'].value
            name7 = active[f'G{4}'].value
            cabinet7 = active[f'A{4}'].value
            name8 = active[f'G{3}'].value
            cabinet8 = active[f'A{3}'].value
            num9 = active[f'H{1}'].value
            name9 = active[f'H{4}'].value
            cabinet9 = active[f'A{4}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num5}. {name5}, каб. {cabinet5};\n \
                            {num5}. {name6}, каб. {cabinet6};\n \
                            {num7}. {name7}, каб. {cabinet7};\n \
                            {num7}. {name8}, каб. {cabinet8};\n \
                            {num9}. {name9}, каб. {cabinet9}; ')
            return lesson
        elif wd == 2:
            dataframe = openpyxl.load_workbook("расписание на среду.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{5}'].value
            cabinet1 = active[f'A{5}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{5}'].value
            cabinet2 = active[f'A{5}'].value
            num3 = active[f'D{1}'].value
            name3 = active[f'D{10}'].value
            cabinet3 = active[f'A{10}'].value
            num4 = active[f'E{1}'].value
            name4 = active[f'E{8}'].value
            cabinet4 = active[f'A{8}'].value
            name5 = active[f'E{9}'].value
            cabinet5 = active[f'A{9}'].value
            num6 = active[f'F{1}'].value
            name6 = active[f'F{8}'].value
            cabinet6 = active[f'A{8}'].value
            name7 = active[f'F{9}'].value
            cabinet7 = active[f'A{9}'].value
            num8 = active[f'G{1}'].value
            name8 = active[f'G{13}'].value
            cabinet8 = active[f'A{13}'].value
            num9 = active[f'H{1}'].value
            name9 = active[f'H{13}'].value
            cabinet9 = active[f'A{13}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num4}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num6}. {name7}, каб. {cabinet7};\n \
                            {num8}. {name8}, каб. {cabinet8};\n \
                            {num9}. {name9}, каб. {cabinet9}; ')
            return lesson
        elif wd == 3:
            dataframe = openpyxl.load_workbook("расписание на четверг.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{14}'].value
            cabinet1 = active[f'A{14}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{14}'].value
            cabinet2 = active[f'A{14}'].value
            num3 = active[f'D{1}'].value
            name3 = active[f'D{13}'].value
            cabinet3 = active[f'A{13}'].value
            num4 = active[f'E{1}'].value
            name4 = active[f'E{13}'].value
            cabinet4 = active[f'A{13}'].value
            num5 = active[f'F{1}'].value
            name5 = active[f'F{10}'].value
            cabinet5 = active[f'A{10}'].value
            num6 = active[f'G{1}'].value
            name6 = active[f'G{10}'].value
            cabinet6 = active[f'A{10}'].value
            num7 = active[f'H{1}'].value
            name7 = active[f'H{5}'].value
            cabinet7 = active[f'A{5}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num5}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num7}. {name7}, каб. {cabinet7};')
            return lesson
        elif wd == 4:
            dataframe = openpyxl.load_workbook("расписание на пятницу.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{13}'].value
            cabinet1 = active[f'A{13}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{13}'].value
            cabinet2 = active[f'A{13}'].value
            num3 = active[f'D{1}'].value
            name3 = active[f'D{13}'].value
            cabinet3 = active[f'A{13}'].value
            num4 = active[f'E{1}'].value
            name4 = active[f'E{2}'].value
            cabinet4 = active[f'A{2}'].value
            num5 = active[f'F{1}'].value
            name5 = active[f'F{2}'].value
            cabinet5 = active[f'A{2}'].value
            num6 = active[f'G{1}'].value
            name6 = active[f'G{5}'].value
            cabinet6 = active[f'A{5}'].value
            num7 = active[f'H{1}'].value
            name7 = active[f'H{5}'].value
            cabinet7 = active[f'A{5}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num5}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num7}. {name7}, каб. {cabinet7};')
            return lesson
    elif user_class == "10И":
        if wd == 0:
            dataframe = openpyxl.load_workbook("расписание на понедельник.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{3}'].value
            cabinet1 = active[f'A{3}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{13}'].value
            cabinet2 = active[f'A{13}'].value
            num3 = active[f'D{1}'].value
            name3 = active[f'D{12}'].value
            cabinet3 = active[f'A{12}'].value
            num4 = active[f'E{1}'].value
            name4 = active[f'E{18}'].value
            cabinet4 = active[f'A{18}'].value
            num5 = active[f'F{1}'].value
            name5 = active[f'F{14}'].value
            cabinet5 = active[f'A{14}'].value
            num6 = active[f'G{1}'].value
            name6 = active[f'G{15}'].value
            cabinet6 = active[f'A{15}'].value
            num7 = active[f'H{1}'].value
            name7 = active[f'H{15}'].value
            cabinet7 = active[f'A{15}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num5}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num7}. {name7}, каб. {cabinet7};')
            return lesson
        elif wd == 1:
            dataframe = openpyxl.load_workbook("расписание на вторник.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{13}'].value
            cabinet1 = active[f'A{13}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{13}'].value
            cabinet2 = active[f'A{13}'].value
            num3 = active[f'D{1}'].value
            name3 = active[f'D{10}'].value
            cabinet3 = active[f'A{10}'].value
            num4 = active[f'E{1}'].value
            name4 = active[f'E{16}'].value
            cabinet4 = active[f'A{16}'].value
            num5 = active[f'F{1}'].value
            name5 = active[f'F{11}'].value
            cabinet5 = active[f'A{11}'].value
            num6 = active[f'G{1}'].value
            name6 = active[f'G{14}'].value
            cabinet6 = active[f'A{14}'].value
            num7 = active[f'H{1}'].value
            name7 = active[f'H{12}'].value
            cabinet7 = active[f'A{12}'].value
            num8 = active[f'I{1}'].value
            name8 = active[f'I{3}'].value
            cabinet8 = active[f'A{3}'].value
            name9 = active[f'I{4}'].value
            cabinet9 = active[f'A{4}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num5}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num7}. {name7}, каб. {cabinet7};\n \
                            {num8}. {name8}, каб. {cabinet8};\n \
                            {num8}. {name9}, каб. {cabinet9};')
            return lesson
        elif wd == 2:
            dataframe = openpyxl.load_workbook("расписание на среду.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{12}'].value
            cabinet1 = active[f'A{12}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{12}'].value
            cabinet2 = active[f'A{12}'].value
            num3 = active[f'D{1}'].value
            name3 = active[f'D{14}'].value
            cabinet3 = active[f'A{14}'].value
            num4 = active[f'E{1}'].value
            name4 = active[f'E{3}'].value
            cabinet4 = active[f'A{3}'].value
            name5 = active[f'E{4}'].value
            cabinet5 = active[f'A{4}'].value
            num6 = active[f'F{1}'].value
            name6 = active[f'F{3}'].value
            cabinet6 = active[f'A{3}'].value
            name7 = active[f'F{4}'].value
            cabinet7 = active[f'A{4}'].value
            num8 = active[f'G{1}'].value
            name8 = active[f'G{10}'].value
            cabinet8 = active[f'A{10}'].value
            num9 = active[f'H{1}'].value
            name9 = active[f'H{10}'].value
            cabinet9 = active[f'A{10}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num4}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num6}. {name7}, каб. {cabinet7};\n \
                            {num8}. {name8}, каб. {cabinet8};\n \
                            {num9}. {name9}, каб. {cabinet9}; ')
            return lesson
        elif wd == 3:
            dataframe = openpyxl.load_workbook("расписание на четверг.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{13}'].value
            cabinet1 = active[f'A{13}'].value
            num2 = active[f'C{1}'].value
            name2 = active[f'C{13}'].value
            cabinet2 = active[f'A{13}'].value
            num3 = active[f'D{1}'].value
            name3 = active[f'D{12}'].value
            cabinet3 = active[f'A{12}'].value
            num4 = active[f'E{1}'].value
            name4 = active[f'E{12}'].value
            cabinet4 = active[f'A{12}'].value
            num5 = active[f'F{1}'].value
            name5 = active[f'F{13}'].value
            cabinet5 = active[f'A{13}'].value
            num6 = active[f'G{1}'].value
            name6 = active[f'G{9}'].value
            cabinet6 = active[f'A{9}'].value
            num7 = active[f'H{1}'].value
            name7 = active[f'H{9}'].value
            cabinet7 = active[f'A{9}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num2}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num5}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num7}. {name7}, каб. {cabinet7};')
            return lesson
        elif wd == 4:
            dataframe = openpyxl.load_workbook("расписание на пятницу.xlsx")
            active = dataframe.active
            lesson = []
            num1 = active[f'B{1}'].value
            name1 = active[f'B{8}'].value
            cabinet1 = active[f'A{8}'].value
            name2 = active[f'B{9}'].value
            cabinet2 = active[f'A{9}'].value
            num3 = active[f'C{1}'].value
            name3 = active[f'C{14}'].value
            cabinet3 = active[f'A{14}'].value
            num4 = active[f'D{1}'].value
            name4 = active[f'D{8}'].value
            cabinet4 = active[f'A{8}'].value
            num5 = active[f'E{1}'].value
            name5 = active[f'E{13}'].value
            cabinet5 = active[f'A{13}'].value
            num6 = active[f'F{1}'].value
            name6 = active[f'F{13}'].value
            cabinet6 = active[f'A{13}'].value
            num7 = active[f'G{1}'].value
            name7 = active[f'G{11}'].value
            cabinet7 = active[f'A{11}'].value
            num8 = active[f'H{1}'].value
            name8 = active[f'H{13}'].value
            cabinet8 = active[f'A{13}'].value
            lesson.append(f'{num1}. {name1}, каб. {cabinet1};\n \
                            {num1}. {name2}, каб. {cabinet2};\n \
                            {num3}. {name3}, каб. {cabinet3};\n \
                            {num4}. {name4}, каб. {cabinet4};\n \
                            {num5}. {name5}, каб. {cabinet5};\n \
                            {num6}. {name6}, каб. {cabinet6};\n \
                            {num7}. {name7}, каб. {cabinet7};\n \
                            {num8}. {name8}, каб. {cabinet8}; ')
            return lesson


def find_user(id):
    for u in users:
        if u['id'] == id:
            return True
    return False


def get_user_class(id):
    for u in users:
        if u['id'] == id:
            return u['class']


def change_user_class(id, new_class):
    for u in users:
        if u['id'] == id:
            u['class'] = new_class
            save_data()


def send_notification(chat_id):
    print(chat_id)
    wd = datetime.datetime.today().weekday()
    if wd == 0:
        clas = get_user_class(chat_id)
        lesson = get_list_of_lessons(clas)
        send_lessons(lesson, chat_id)
    elif wd == 1:
        clas = get_user_class(chat_id)
        lesson = get_list_of_lessons(clas)
        send_lessons(lesson, chat_id)
    elif wd == 2:
        clas = get_user_class(chat_id)
        lesson = get_list_of_lessons(clas)
        send_lessons(lesson, chat_id)
    elif wd == 3:
        clas = get_user_class(chat_id)
        lesson = get_list_of_lessons(clas)
        send_lessons(lesson, chat_id)
    elif wd == 4:
        clas = get_user_class(chat_id)
        lesson = get_list_of_lessons(clas)
        send_lessons(lesson, chat_id)


def get_day_two(clas):
    wd = datetime.datetime.today().weekday()
    if clas == "10Т":
        if wd == 0:
            lesson = send_lesson_10t(2)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10t(2)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10t(2)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10t(2)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10t(2)
            return lesson
    if clas == "10И":
        if wd == 0:
            lesson = send_lesson_10i(2)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10i(2)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10i(2)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10i(2)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10i(2)
            return lesson


def send_day_second(chat_id):
    print(chat_id)
    clas = get_user_class(chat_id)
    lesson = get_day_two(clas)
    send_lesson(lesson, chat_id)


def get_day_three(clas):
    wd = datetime.datetime.today().weekday()
    if clas == "10Т":
        if wd == 0:
            lesson = send_lesson_10t(3)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10t(3)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10t(3)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10t(3)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10t(3)
            return lesson
    if clas == "10И":
        if wd == 0:
            lesson = send_lesson_10i(3)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10i(3)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10i(3)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10i(3)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10i(3)
            return lesson


def send_day_third(chat_id):
    print(chat_id)
    clas = get_user_class(chat_id)
    lesson = get_day_three(clas)
    send_lesson(lesson, chat_id)


def get_day_four(clas):
    wd = datetime.datetime.today().weekday()
    if clas == "10Т":
        if wd == 0:
            lesson = send_lesson_10t(4)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10t(4)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10t(4)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10t(4)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10t(4)
            return lesson
    if clas == "10И":
        if wd == 0:
            lesson = send_lesson_10i(4)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10i(4)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10i(4)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10i(4)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10i(4)
            return lesson


def send_day_forth(chat_id):
    print(chat_id)
    clas = get_user_class(chat_id)
    lesson = get_day_four(clas)
    send_lesson(lesson, chat_id)


def get_day_five(clas):
    wd = datetime.datetime.today().weekday()
    if clas == "10Т":
        if wd == 0:
            lesson = send_lesson_10t(5)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10t(5)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10t(5)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10t(5)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10t(5)
            return lesson
    if clas == "10И":
        if wd == 0:
            lesson = send_lesson_10i(5)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10i(5)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10i(5)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10i(5)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10i(5)
            return lesson


def send_day_fifth(chat_id):
    print(chat_id)
    clas = get_user_class(chat_id)
    lesson = get_day_five(clas)
    send_lesson(lesson, chat_id)


def get_day_six(clas):
    wd = datetime.datetime.today().weekday()
    if clas == "10Т":
        if wd == 0:
            lesson = send_lesson_10t(6)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10t(6)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10t(6)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10t(6)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10t(6)
            return lesson
    if clas == "10И":
        if wd == 0:
            lesson = send_lesson_10i(6)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10i(6)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10i(6)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10i(6)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10i(6)
            return lesson


def send_day_sixth(chat_id):
    print(chat_id)
    clas = get_user_class(chat_id)
    lesson = get_day_six(clas)
    send_lesson(lesson, chat_id)


def get_day_seven(clas):
    wd = datetime.datetime.today().weekday()
    if clas == "10Т":
        if wd == 0:
            lesson = send_lesson_10t(7)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10t(7)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10t(7)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10t(7)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10t(7)
            return lesson
    if clas == "10И":
        if wd == 0:
            lesson = send_lesson_10i(7)
            return lesson
        elif wd == 1:
            lesson = send_lesson_10i(7)
            return lesson
        elif wd == 2:
            lesson = send_lesson_10i(7)
            return lesson
        elif wd == 3:
            lesson = send_lesson_10i(7)
            return lesson
        elif wd == 4:
            lesson = send_lesson_10i(7)
            return lesson


def send_day_seventh(chat_id):
    print(chat_id)
    clas = get_user_class(chat_id)
    lesson = get_day_seven(clas)
    send_lesson(lesson, chat_id)


def get_day_eight(clas):
    wd = datetime.datetime.today().weekday()
    if clas == "10Т":
        if wd == 0:
            lesson = send_lesson_10t(8)
            return lesson
    elif clas == "10И":
        if wd == 1:
            lesson = send_lesson_10i(8)
            return lesson


def send_day_eighth(chat_id):
    print(chat_id)
    clas = get_user_class(chat_id)
    lesson = get_day_eight(clas)
    send_lesson(lesson, chat_id)


def send_lesson(lesson, chat_id):
    bot.send_message(chat_id, lesson)


def run_bot():
    while True:
        try:
            bot.polling(none_stop=True)
        except:
            pass


def run_schedule():
    while True:
        schedule.run_pending()
        time.sleep(1)


if __name__ == '__main__':
    import threading

    threading.Thread(target=run_bot).start()
    threading.Thread(target=run_schedule).start()
